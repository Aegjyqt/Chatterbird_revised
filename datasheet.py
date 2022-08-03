import openpyxl
from aiogram.utils.markdown import hcode


class Datasheet:
    """ Governs essential data fom a specific sheet """
    _all_categories = []  # как реализовать это как set?
    _translations_entire_dict = {}
    """ The following attributes are for specific columns in .xlsx file sheets: """
    _name_column = 0  # еще раз проверить, кажется не используется вообще
    _ids_column = 1
    _relevance_column = 2
    _keys_column = 3
    _items_column = 4

    def __init__(self) -> None:
        self.category_dict = {}
        self.sheet_data = []  # как тут можно применить Namedtuple? и целесообразно ли?

    def specify_sheet_data(self, wb: str, name: str) -> (str, str):
        """ Sets workbook and sheet name for a specific class instance """
        this_sheet_data = [wb, name]  # name также используется в user_input, подумать как сделать лучше
        self.sheet_data = this_sheet_data
        self._all_categories.append(self)

    def get_ids(self) -> list:
        """ Gets category IDs from a specific sheet, appends the complete list of categories """
        wb = openpyxl.load_workbook(self.sheet_data[0])
        sheet = wb[self.sheet_data[1]]
        category_ids = []
        for row in range(2, sheet.max_row):  # почему не получается с просто range(sheet.max_row)?
            category_ids.append(sheet[row][self._ids_column].value)
        return category_ids

    def get_category_dict(self) -> None:
        """ Gets keys and items for a category dictionary, updates complete translation dictionary """
        wb = openpyxl.load_workbook(self.sheet_data[0])
        sheet = wb[self.sheet_data[1]]
        keys = []
        items = []
        for i in range(2, sheet.max_row):
            temp = list(sheet[i][self._keys_column].value.lower())
            for n in temp:
                if n == '\"':
                    temp.remove(n)
            key = ''.join(temp)
            keys.append(key)
            items.append(sheet[i][self._items_column].value)
            final_item = [hcode(items[keys.index(key)]), f"актуальность: {sheet[i][self._relevance_column].value}"]
            self.category_dict[key] = final_item
        self._translations_entire_dict.update(self.category_dict)

    def get_all_categories(self) -> list:
        """ Meant for getting the _all_categories attribute outside of class """
        return self._all_categories

    def get_translations_entire_dict(self) -> dict:
        """ Meant for getting the _translations_entire_dict attribute outside of class """
        return self._translations_entire_dict


departments = Datasheet()
departments.specify_sheet_data('lists_for_chatterbird.xlsx', 'управление')
departments.get_ids()  # эти вызовы можно как-то сократить? => попробуй сделать класс
departments.get_category_dict()
faculties = Datasheet()
faculties.specify_sheet_data('lists_for_chatterbird.xlsx', 'факультет')
faculties.get_ids()
faculties.get_category_dict()

